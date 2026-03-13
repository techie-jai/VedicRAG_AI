import openpyxl
import sqlite3
import csv
import io
from pathlib import Path

def setup_scriptures_database():
    """
    Read the 200 Scriptures Title.xlsx file (with CSV data) and create a SQLite database
    """
    excel_path = Path('dharmaganj/200 Scriptures Title.xlsx')
    db_path = Path('scriptures.db')
    
    # Load Excel file
    print(f"Loading Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    print(f"Sheet name: {ws.title}")
    print(f"Total rows: {ws.max_row}")
    
    # Extract all data from Excel and parse as CSV
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        if row and row[0]:
            all_rows.append(row[0])
    
    # Parse CSV data
    csv_reader = csv.reader(all_rows)
    headers = next(csv_reader)
    
    print(f"\nColumn headers: {headers}")
    print(f"Number of columns: {len(headers)}\n")
    
    # Display first few rows to verify structure
    print("First 5 data rows:")
    data_rows = list(csv_reader)
    for i, row in enumerate(data_rows[:5], 1):
        print(f"Row {i}: {row}")
    
    # Create SQLite database
    print(f"\nCreating SQLite database: {db_path}")
    
    # Remove existing database if it exists
    if db_path.exists():
        db_path.unlink()
        print(f"Removed existing database")
    
    conn = sqlite3.connect(str(db_path))
    cursor = conn.cursor()
    
    # Create table with dynamic columns
    # Sanitize column names (remove special characters, spaces)
    sanitized_headers = []
    for header in headers:
        sanitized = header.strip().replace(' ', '_').replace('-', '_').lower()
        sanitized_headers.append(sanitized)
    
    # Create table schema
    columns_def = ', '.join([f"{col} TEXT" for col in sanitized_headers])
    create_table_sql = f"CREATE TABLE scriptures (id INTEGER PRIMARY KEY AUTOINCREMENT, {columns_def})"
    
    print(f"\nTable schema:\n{create_table_sql}\n")
    cursor.execute(create_table_sql)
    
    # Insert data into database
    insert_sql = f"INSERT INTO scriptures ({', '.join(sanitized_headers)}) VALUES ({', '.join(['?' for _ in sanitized_headers])})"
    
    row_count = 0
    for row in data_rows:
        if any(row):
            cursor.execute(insert_sql, row)
            row_count += 1
    
    conn.commit()
    
    # Verify insertion
    cursor.execute("SELECT COUNT(*) FROM scriptures")
    total_records = cursor.fetchone()[0]
    
    print(f"Successfully inserted {total_records} records into the database")
    
    # Display sample records
    print("\nSample records from database:")
    cursor.execute(f"SELECT * FROM scriptures LIMIT 3")
    for record in cursor.fetchall():
        print(f"\nID: {record[0]}")
        for i, header in enumerate(sanitized_headers):
            print(f"  {header}: {record[i+1][:100] if record[i+1] else 'N/A'}...")
    
    conn.close()
    print(f"\nDatabase created successfully at: {db_path}")

if __name__ == "__main__":
    setup_scriptures_database()
