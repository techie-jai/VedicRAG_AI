import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

def get_file_size(filepath):
    """Get file size in KB"""
    try:
        size_bytes = os.path.getsize(filepath)
        return size_bytes / 1024
    except:
        return 0

def get_scripture_description(filename, domain):
    """Get description for a scripture based on filename and domain"""
    descriptions = {
        # Vedas
        "vedic_corpus": "Vedic texts - Rig, Sama, Yajur, Atharva Vedas with hymns and rituals",
        "Rig Veda": "Oldest Veda with 1,028 hymns praising deities and describing cosmology",
        "Sama Veda": "Veda of Melodies - hymns arranged for liturgical chanting",
        "Yajur Veda": "Veda of Sacrifices - detailed instructions for Vedic rituals",
        "Atharva Veda": "Veda of Incantations - spells, charms, and metaphysical teachings",
        
        # Upanishads
        "upanishad": "Philosophical treatises on Brahman and Atman",
        "Upanishads": "Foundation of Vedantic philosophy exploring ultimate reality",
        
        # Bhagavad Gita
        "Bhagavad Gita": "Sacred dialogue between Krishna and Arjuna on dharma and liberation",
        "srimadbhagavadgita": "Bhagavad Gita - teachings on duty, devotion, and paths to liberation",
        
        # Puranas
        "purana": "Mythological narratives on cosmology and divine principles",
        "Puranas": "Post-Vedic texts with mythology, cosmology, and spiritual teachings",
        
        # Epics
        "itihasa": "Great Epics - Ramayana and Mahabharata with philosophical teachings",
        "Mahabharata": "World's longest epic on the Kurukshetra war and human nature",
        "Ramayana": "Epic tale of Prince Rama's exile teaching dharma and righteousness",
        
        # Medicine (Ayurveda)
        "cikitsavidya": "Ayurvedic medical treatises on healing and herbal remedies",
        "carakasamhita": "Charaka Samhita - foundational Ayurvedic medical text",
        "susrutasamhita": "Sushruta Samhita - surgical treatises and medical procedures",
        "astangahrdayasamhita": "Ashtanga Hridaya - comprehensive Ayurvedic medicine guide",
        
        # Logic & Epistemology
        "nyaya_pramana": "Logical treatises on epistemology and valid knowledge",
        "nyayamanjari": "Compendium of Nyaya logic and philosophical reasoning",
        "pramanavarttika": "Buddhist epistemology and theory of valid knowledge",
        
        # Grammar
        "vyakarana": "Sanskrit grammar and linguistic analysis",
        "bhartrhari-vakyapadiya": "Philosophy of language and sentence meaning",
        
        # Political Science
        "arthashastra": "Kautilya's treatise on statecraft and political economy",
        "kautalyarthasastra": "Ancient text on governance, economics, and diplomacy",
        
        # Classical Poetry & Drama
        "kavya": "Classical Sanskrit poetry and dramatic literature",
        "bana-kadambari": "Romantic narrative prose by Bana",
        "asvaghosa-buddhacarita": "Life of Buddha in poetic form",
        
        # Sutras & Philosophical Works
        "sutra": "Aphoristic philosophical texts and commentaries",
        "patanjalayogasastra": "Yoga Sutras - foundational text on Raja Yoga",
        "astavakragita": "Philosophical teachings on non-dualism",
        
        # Modern & Contemporary
        "vividha": "Miscellaneous texts including modern and contemporary works",
    }
    
    # Check various keys for description
    for key, desc in descriptions.items():
        if key.lower() in filename.lower() or key.lower() in domain.lower():
            return desc
    
    return f"Text from {domain} domain"

def scan_dharmaganj():
    """Scan dharmaganj folder and collect all scripture information"""
    
    dharmaganj_path = "dharmaganj"
    scriptures = []
    
    # Walk through all directories
    for root, dirs, files in os.walk(dharmaganj_path):
        for file in files:
            if file.endswith(('.txt', '.xml', '.json')) and file != 'master_catalog.json':
                filepath = os.path.join(root, file)
                rel_path = filepath.replace('\\', '/')
                
                # Parse the path to extract building, domain, and file info
                parts = rel_path.split('/')
                
                if len(parts) >= 3:
                    building = parts[1]
                    domain = parts[2] if len(parts) > 2 else "unknown"
                    
                    # Determine language and format
                    file_ext = os.path.splitext(file)[1]
                    if file_ext == '.xml':
                        file_format = "XML"
                        language = "Sanskrit"
                    elif file_ext == '.txt':
                        file_format = "TXT"
                        language = "Sanskrit"
                    elif file_ext == '.json':
                        file_format = "JSON"
                        language = "Sanskrit"
                    else:
                        file_format = file_ext.upper()
                        language = "Unknown"
                    
                    file_size = get_file_size(filepath)
                    description = get_scripture_description(file, domain)
                    
                    scripture_info = {
                        "building": building,
                        "domain": domain,
                        "description": description,
                        "filename": file,
                        "format": file_format,
                        "language": language,
                        "size_kb": file_size,
                        "path": rel_path
                    }
                    
                    scriptures.append(scripture_info)
    
    return scriptures

def create_excel(scriptures):
    """Create Excel workbook with scripture data"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Scripture Index"
    
    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Building colors
    building_colors = {
        "ratnodadhi": "E7E6E6",
        "ratnasagara": "FFF2CC",
        "ratnaranjaka": "E2EFDA",
        "bagdevibhandar": "FCE4D6"
    }
    
    # Headers
    headers = ["Building", "Domain", "Description", "Filename", "Format", "Language", "Size (KB)", "Path"]
    ws.append(headers)
    
    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    
    # Add data rows
    for scripture in sorted(scriptures, key=lambda x: (x['building'], x['domain'], x['filename'])):
        building = scripture['building']
        row_data = [
            building,
            scripture['domain'],
            scripture['description'],
            scripture['filename'],
            scripture['format'],
            scripture['language'],
            f"{scripture['size_kb']:.1f}",
            scripture['path']
        ]
        ws.append(row_data)
        
        # Style data row
        row_num = ws.max_row
        building_color = building_colors.get(building, "FFFFFF")
        fill = PatternFill(start_color=building_color, end_color=building_color, fill_type="solid")
        
        for cell in ws[row_num]:
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15  # Building
    ws.column_dimensions['B'].width = 20  # Domain
    ws.column_dimensions['C'].width = 50  # Description
    ws.column_dimensions['D'].width = 35  # Filename
    ws.column_dimensions['E'].width = 10  # Format
    ws.column_dimensions['F'].width = 12  # Language
    ws.column_dimensions['G'].width = 12  # Size
    ws.column_dimensions['H'].width = 50  # Path
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Create summary sheet
    summary_ws = wb.create_sheet("Summary")
    
    # Count by building
    building_counts = defaultdict(int)
    domain_counts = defaultdict(int)
    format_counts = defaultdict(int)
    
    for scripture in scriptures:
        building_counts[scripture['building']] += 1
        domain_counts[scripture['domain']] += 1
        format_counts[scripture['format']] += 1
    
    # Summary header
    summary_ws['A1'] = "SCRIPTURE INDEX SUMMARY"
    summary_ws['A1'].font = Font(bold=True, size=14)
    
    # Building summary
    summary_ws['A3'] = "Texts by Building"
    summary_ws['A3'].font = Font(bold=True, size=12)
    summary_ws['A4'] = "Building"
    summary_ws['B4'] = "Count"
    
    for cell in [summary_ws['A4'], summary_ws['B4']]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    row = 5
    for building in sorted(building_counts.keys()):
        summary_ws[f'A{row}'] = building
        summary_ws[f'B{row}'] = building_counts[building]
        row += 1
    
    # Domain summary
    summary_ws['A12'] = "Texts by Domain"
    summary_ws['A12'].font = Font(bold=True, size=12)
    summary_ws['A13'] = "Domain"
    summary_ws['B13'] = "Count"
    
    for cell in [summary_ws['A13'], summary_ws['B13']]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    row = 14
    for domain in sorted(domain_counts.keys()):
        summary_ws[f'A{row}'] = domain
        summary_ws[f'B{row}'] = domain_counts[domain]
        row += 1
    
    # Format summary
    summary_ws['D3'] = "Texts by Format"
    summary_ws['D3'].font = Font(bold=True, size=12)
    summary_ws['D4'] = "Format"
    summary_ws['E4'] = "Count"
    
    for cell in [summary_ws['D4'], summary_ws['E4']]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    row = 5
    for fmt in sorted(format_counts.keys()):
        summary_ws[f'D{row}'] = fmt
        summary_ws[f'E{row}'] = format_counts[fmt]
        row += 1
    
    # Total count
    summary_ws['A20'] = "Total Texts"
    summary_ws['A20'].font = Font(bold=True, size=12)
    summary_ws['B20'] = len(scriptures)
    summary_ws['B20'].font = Font(bold=True, size=12)
    
    # Adjust summary sheet columns
    summary_ws.column_dimensions['A'].width = 20
    summary_ws.column_dimensions['B'].width = 12
    summary_ws.column_dimensions['D'].width = 20
    summary_ws.column_dimensions['E'].width = 12
    
    return wb

def main():
    print("Scanning dharmaganj folder...")
    scriptures = scan_dharmaganj()
    
    print(f"Found {len(scriptures)} texts")
    print("Creating Excel workbook...")
    
    wb = create_excel(scriptures)
    
    output_file = "SCRIPTURE_INDEX.xlsx"
    wb.save(output_file)
    
    print(f"\n✅ Excel file created successfully!")
    print(f"📊 Output file: {output_file}")
    print(f"\n📈 Statistics:")
    print(f"   Total texts: {len(scriptures)}")
    
    # Count by building
    building_counts = defaultdict(int)
    for scripture in scriptures:
        building_counts[scripture['building']] += 1
    
    for building in sorted(building_counts.keys()):
        print(f"   {building}: {building_counts[building]} texts")

if __name__ == "__main__":
    main()
